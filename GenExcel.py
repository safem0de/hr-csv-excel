import json
import os
import pandas as pd

from datetime import datetime
from openpyxl import load_workbook

class GenExcel:
    def __init__(self, csv_path, excel_path, config_path="config.json"):
        self.csv_path = csv_path
        self.excel_path = excel_path
        self.config_path  = config_path
        self.config = self._load_config()

    def _load_config(self):
        if not os.path.exists(self.config_path):
            raise FileNotFoundError(f"ไม่พบ config: {self.config_path}")
        with open(self.config_path, encoding="utf-8") as f:
            return json.load(f)

    def readconfig(self, section, key=None):
        try:
            if key is None:
                return self.config[section]
            else :
                return self.config[section][key]
        except KeyError as e:
            raise KeyError(f"ไม่พบ key: {e}")
        
    def create_sheetname_from_csv(self, csv_data):
        try:
            _col = self.readconfig("create_sheetname","from_column")
            # csv_data = pd.read_csv(self.csv_path)
            sheet_name = csv_data[_col].str.replace(r"\s+", " ", regex=True).str.strip().dropna().tolist()
            return sheet_name
        except Exception as e:
            raise Exception(f"create_sheetname_from_csv พบข้อผิดพลาด : {e}")

    def generateExcel(self):
        try:
            now = datetime.now()
            timestamp_str = now.strftime("%d-%m-%Y_%H%M%S")

            xlsx_path = self.readconfig("excel_path")
            wb = load_workbook(xlsx_path)
            ws = wb.active
            template_sheet = wb["BTL"]

            _csv_data = pd.read_csv(self.csv_path)
            _sheet_name = self.create_sheetname_from_csv(_csv_data)
            

            # วนลูปชื่อผู้สมัครพร้อมดัชนี
            for i, name in enumerate(_sheet_name):
                # สร้างชื่อชีตแบบปลอดภัย
                safe_name = "".join(c for c in name if c not in r'[]:*?/\\')[:31]

                # สร้างสำเนา Template
                new_sheet = wb.copy_worksheet(template_sheet)
                new_sheet.title = safe_name

                # ดึงแถวข้อมูลของผู้สมัครจาก CSV (ตามลำดับ)
                row_data = _csv_data.iloc[i]

                # วนลูป config.filled_data เพื่อเติมข้อมูลใน cell
                for label, mapping in self.config["filled_data"].items():
                    from_column = mapping["from"]         # ชื่อ column ใน csv
                    to_cell = mapping["to"]               # cell ใน excel

                    if from_column in row_data:
                        value = row_data[from_column]
                        new_sheet[to_cell] = value
                    else:
                        print(f"Warning: Column '{from_column}' not found in CSV")
                
                # ใส่คะแนนประเมิน (✓ ลง cell)
                for topic, mapping in self.config["scoring"].items():
                    score_str = str(row_data.get(mapping["from"], "")).strip()
                    row = mapping["row"]
                    if score_str in self.config["data_point"]:
                        col_letter = self.config["data_point"][score_str]["column"]
                        cell = f"{col_letter}{row}"
                        new_sheet[cell] = "✓"

            # บันทึกไฟล์
            wb.save(f"{timestamp_str}_interview_filled.xlsx")
        except Exception as e:
            raise Exception(f"generateExcel พบข้อผิดพลาด : {e}")