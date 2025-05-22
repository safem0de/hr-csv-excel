import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pathlib import Path

# Step 1: Read CSV (simulate CSV content)
csv_data = """Name,Score
Alice,85
Bob,92
Charlie,78
"""

# Save mock CSV file
csv_path = Path("mock_data.csv")
csv_path.write_text(csv_data)

# Read CSV with pandas
df = pd.read_csv(csv_path)

# Step 2: Create new Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Report"

# Step 3: Write headers manually to Excel
ws["B1"] = "Name"
ws["C1"] = "Score"

# Step 4: Write data to specific cells starting at B2
start_row = 2
for i, row in df.iterrows():
    ws[f"B{start_row + i}"] = row["Name"]
    ws[f"C{start_row + i}"] = row["Score"]

# Save Excel file
excel_path = Path("output.xlsx")
wb.save(excel_path)