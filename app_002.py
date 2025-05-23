import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import json
from datetime import datetime

now = datetime.now()
timestamp_str = now.strftime("%d-%m-%Y_%H%M%S")

# โหลด config.json
with open("config.json", encoding="utf-8") as f:
    config = json.load(f)

# ดึงชื่อคอลัมน์
_col = config["create_sheetname"]["from_column"]
print(_col)

csv_path = "interview_data.csv"
xls_path = "interview_format.xlsx"

csv_data = pd.read_csv(csv_path)
print(csv_data.head())

candidate = csv_data[_col].str.replace(r"\s+", " ", regex=True).str.strip().dropna().tolist()
print(candidate)

df = pd.read_excel(xls_path)
print(df.head())

wb = load_workbook(xls_path) # เปิด Excel Workbook
ws = wb.active # Active
# ชื่อชีตต้นฉบับ
template_sheet = wb["BTL"]

# วนลูปชื่อผู้สมัครพร้อมดัชนี
for i, name in enumerate(candidate):
    # สร้างชื่อชีตแบบปลอดภัย
    safe_name = "".join(c for c in name if c not in r'[]:*?/\\')[:31]

    # สร้างสำเนา Template
    new_sheet = wb.copy_worksheet(template_sheet)
    new_sheet.title = safe_name

    # ดึงแถวข้อมูลของผู้สมัครจาก CSV (ตามลำดับ)
    row_data = csv_data.iloc[i]

    # วนลูป config.filled_data เพื่อเติมข้อมูลใน cell
    for label, mapping in config["filled_data"].items():
        from_column = mapping["from"]         # ชื่อ column ใน csv
        to_cell = mapping["to"]               # cell ใน excel

        if from_column in row_data:
            value = row_data[from_column]
            new_sheet[to_cell] = value
        else:
            print(f"Warning: Column '{from_column}' not found in CSV")
    
    # ใส่คะแนนประเมิน (✓ ลง cell)
    for topic, mapping in config["scoring"].items():
        score_str = str(row_data.get(mapping["from"], "")).strip()
        row = mapping["row"]
        if score_str in config["data_point"]:
            col_letter = config["data_point"][score_str]["column"]
            cell = f"{col_letter}{row}"
            new_sheet[cell] = "✓"

# บันทึกไฟล์
wb.save(f"{timestamp_str}_interview_filled.xlsx")